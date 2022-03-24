from openpyxl import *
import os
# filename = '20220312_eth_uniswap_10.xlsx'


class Excel:

    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    def getColValues(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data


def getAddress(filename):
    if len(filename) == 0:
        print('未指定地址文件')
        return
    # 用户地址路径，以xlsx格式保存
    file = getFilePath(filename)
    address_list = Excel(file).getColValues(1)
    return address_list

def getFilePath(filename):
    pwd=os.getcwd()
    absPath = os.path.abspath(os.path.dirname(pwd))
    filePath = os.path.join(absPath, 'address/' + filename)
    return filePath

def getSeedPhrase(filename, address):
    input_address = address
    if len(filename) == 0:
        print('未指定地址文件')
        return
    # 用户助记词路径，以xlsx格式保存，该路径由用户自行修改
    file = getFilePath(filename)
    address_list = Excel(file).getColValues(1)
    mnemonic_list = Excel(file).getColValues(3)

    for i in range(1, len(address_list)):
        if input_address == address_list[i]:
            num = i
            return mnemonic_list[num]


def getSeedPhraseV2(filepath):
    # 用户助记词路径，以xlsx格式保存，该路径由用户提供
    # example:'/Users/luoye/Downloads/TestNetwork/20220317_eth_zkSync_muteSwitch_100.xlsx'
    file = filepath
    address_list = Excel(file).getColValues(1)
    mnemonic_list = Excel(file).getColValues(3)
    wallet = [address_list, mnemonic_list]
    return wallet



